import openpyxl
import json
import os
import re
import time 

# --- Configuration ---
EXCEL_FILE = 'viva_links_fixed.xlsx'
IMAGE_FOLDER_NAME = 'downloaded_images' 
OUTPUT_SCRIPT_FILE = 'script.js'
OUTPUT_HTML_FILE = 'index.html'
OUTPUT_CSS_FILE = 'style.css'
SHEETS_TO_PROCESS = {
    'anat': 'Anat',
    'path': 'Path',
    'physio': 'Physio',
    'pharm': 'Pharm',
    'cbb': 'CBB'
}
SHEET_COLORS = {
    'anat': '#3498db',
    'path': '#e74c3c',
    'physio': '#2ecc71',
    'pharm': '#9b59b6',
    'cbb': '#f39c12'
}

def generate_web_data():
    """Reads the Excel file and extracts all data."""
    print(f"Reading data from '{EXCEL_FILE}'...")
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ Error: Excel file not found at '{EXCEL_FILE}'")
        return None

    workbook = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
    workbook_data_only = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    
    all_subject_data = {}

    for key, sheet_name in SHEETS_TO_PROCESS.items():
        if sheet_name not in workbook.sheetnames:
            print(f"⚠️ Warning: Sheet '{sheet_name}' not found. Skipping.")
            continue
        
        worksheet = workbook[sheet_name]
        worksheet_data_only = workbook_data_only[sheet_name]
        
        headers = [cell.value for cell in worksheet[2]]
        sheet_sections = []
        
        for i, header_name in enumerate(headers):
            if header_name and isinstance(header_name, str):
                sheet_sections.append({ "header": header_name.strip(), "col_index": i, "items": [] })
        
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=3), start=3):
            for section in sheet_sections:
                desc_col_index = section['col_index'] + 1
                status_col_index = section['col_index'] + 2
                
                if len(row) > status_col_index:
                    desc_cell = row[desc_col_index] 
                    status_cell = row[status_col_index]

                    if desc_cell.value:
                        item_data = { "desc": "", "status": str(status_cell.value) if status_cell.value else "Pending", "hyperlink": "" }
                        
                        if desc_cell.hyperlink:
                            item_data["desc"] = str(desc_cell.value)
                            item_data["hyperlink"] = desc_cell.hyperlink.target.replace("\\", "/")
                        elif isinstance(desc_cell.value, str) and desc_cell.value.strip().upper().startswith('=HYPERLINK'):
                            match = re.search(r'=HYPERLINK\("([^"]+)"(?:,\s*"([^"]+)")?\)', desc_cell.value, re.IGNORECASE)
                            if match:
                                link_target = match.group(1)
                                friendly_name_cell = worksheet_data_only.cell(row=row_idx, column=desc_col_index + 1)
                                item_data["desc"] = str(friendly_name_cell.value)
                                item_data["hyperlink"] = link_target.replace("\\", "/")
                            else:
                                friendly_name_cell = worksheet_data_only.cell(row=row_idx, column=desc_col_index + 1)
                                item_data["desc"] = str(friendly_name_cell.value)
                        else:
                            item_data["desc"] = str(desc_cell.value)
                        
                        if item_data["hyperlink"]:
                            item_data["hyperlink"] = item_data["hyperlink"].replace("downloaded images/", f"{IMAGE_FOLDER_NAME}/")
                        
                        section["items"].append(item_data)
        
        all_subject_data[key] = {
            "title": sheet_name,
            "color": SHEET_COLORS.get(key, '#7f8c8d'),
            "data": [s for s in sheet_sections if s["items"]]
        }
    print("✅ Data extraction complete.")
    return all_subject_data

def write_html_file():
    """Writes the HTML file, forcing a desktop viewport on all devices."""
    cache_buster = int(time.time())
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=1280">
    <title>ED Primary VIVA and MCQ Resources</title>
    <link rel="stylesheet" href="style.css?v={cache_buster}">
</head>
<body>
    <div id="timer-container">
        <div class="timer-module">
            <div id="timer-display-2min" class="timer-display">02:00</div>
            <div class="timer-buttons">
                <button id="timer-control-2min">Start</button>
                <button id="timer-reset-2min">Reset</button>
            </div>
        </div>
        <div class="timer-module">
            <div id="timer-display-10min" class="timer-display">10:00</div>
            <div class="timer-buttons">
                <button id="timer-control-10min">Start</button>
                <button id="timer-reset-10min">Reset</button>
            </div>
        </div>
    </div>

    <header><h1>ED Primary VIVA and MCQ Resources</h1><nav><button class="tab-btn active" data-tab="dashboard">Dashboard</button><button class="tab-btn" data-tab="anat">Anat</button><button class="tab-btn" data-tab="path">Path</button><button class="tab-btn" data-tab="physio">Physio</button><button class="tab-btn" data-tab="pharm">Pharm</button><button class="tab-btn" data-tab="cbb">CBB</button></nav></header>
    <main>
        <div id="dashboard" class="tab-content active">
            <div class="dashboard-layout">
                <div class="left-column">
                    <h2>Overall Progress</h2>
                    <div class="progress-container"><div class="progress-bar" id="total-progress-bar"></div><span class="progress-label" id="total-progress-label"></span></div>
                    <div class="subject-progress-grid"><div id="anat-progress-card" class="progress-card"></div><div id="path-progress-card" class="progress-card"></div><div id="physio-progress-card" class="progress-card"></div><div id="pharm-progress-card" class="progress-card"></div><div id="cbb-progress-card" class="progress-card"></div></div>
                    <div class="downloads-section">
                        <h2>Downloads & Links</h2>
                        <a href="combined1.pdf" class="download-link" download>Viva Stations Compiled to 2025</a>
                        <a href="viva_links_fixed.xlsx" class="download-link" download>Download Excel Source File</a>
                        <a href="https://notebooklm.google.com/notebook/e1cdd0b3-3f57-452b-84e3-bc3fd476d80c?authuser=3" class="download-link" target="_blank">NotebookLM Viva</a>
                    </div>
                </div>
                <div class="right-column">
                    <div class="mcq-section">
                        <h2>MCQs</h2>
                        <div class="mcq-category mcq-physio">
                            <h3>Physio</h3>
                            <a href="IM Physio.htm" class="mcq-link" target="_blank">IM Physio</a>
                            <a href="ME Physio.htm" class="mcq-link" target="_blank">ME Physio</a>
                            <a href="Geelong Physio.htm" class="mcq-link" target="_blank">Geelong Physio</a>
                        </div>
                        <div class="mcq-category mcq-pharm">
                            <h3>Pharm</h3>
                            <a href="IM Pharm.htm" class="mcq-link" target="_blank">IM Pharm</a>
                            <a href="ME Pharm.htm" class="mcq-link" target="_blank">ME Pharm</a>
                            <a href="Geelong Pharm.htm" class="mcq-link" target="_blank">Geelong Pharm</a>
                        </div>
                        <div class="mcq-category mcq-anat">
                            <h3>Anat</h3>
                            <a href="IM Anat.htm" class="mcq-link" target="_blank">IM Anat</a>
                            <a href="ME Anat.htm" class="mcq-link" target="_blank">ME Anat</a>
                            <a href="Geelong Anat.htm" class="mcq-link" target="_blank">Geelong Anat</a>
                        </div>
                        <div class="mcq-category mcq-path">
                            <h3>Path</h3>
                            <a href="IM Path.htm" class="mcq-link" target="_blank">IM Path</a>
                            <a href="ME Path.htm" class="mcq-link" target="_blank">ME Path</a>
                            <a href="Geelong Path.htm" class="mcq-link" target="_blank">Geelong Path</a>
                        </div>
                        <div class="mcq-category mcq-random">
                            <h3>Random Collection</h3>
                            <a href="Anatmerged.html" class="mcq-link" target="_blank">Anat Merged</a>
                            <a href="pathmerged.html" class="mcq-link" target="_blank">Path Merged</a>
                            <a href="pharmmerged.html" class="mcq-link" target="_blank">Pharm Merged</a>
                            <a href="physiomerged.html" class="mcq-link" target="_blank">Physio Merged</a>
                            <a href="emcq1.html" class="mcq-link" target="_blank">EMCQ 1</a>
                            <a href="180mixed.htm" class="mcq-link" target="_blank">180 Mixed</a>
                        </div>
                    </div>
                </div>
            </div>
        </div>
        <div id="anat" class="tab-content"></div><div id="path" class="tab-content"></div><div id="physio" class="tab-content"></div><div id="pharm" class="tab-content"></div><div id="cbb" class="tab-content"></div>
    </main>
    <script src="script.js?v={cache_buster}"></script>
</body>
</html>"""
    with open(OUTPUT_HTML_FILE, 'w', encoding='utf-8') as f: f.write(html_content)
    print(f"✅ Created '{OUTPUT_HTML_FILE}'")

def write_css_file():
    """Writes the CSS file."""
    css_content = """body{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Helvetica,Arial,sans-serif;background-color:#f4f7f9;color:#333;margin:0;padding:20px}header{text-align:center;margin-bottom:20px}h1{color:#2c3e50;font-size:2em}h2{color:#34495e;border-bottom:2px solid #e0e0e0;padding-bottom:10px;margin-top:40px}nav{display:flex;justify-content:center;background-color:#fff;border-radius:8px;padding:5px;box-shadow:0 2px 4px rgba(0,0,0,.1);margin-bottom:20px}.tab-btn{padding:10px 20px;border:none;background-color:transparent;cursor:pointer;font-size:16px;border-radius:6px;transition:background-color .3s,color .3s}.tab-btn.active{background-color:#3498db;color:#fff}.tab-btn:hover:not(.active){background-color:#ecf0f1}.tab-content{display:none;padding:20px;background-color:#fff;border-radius:8px;box-shadow:0 2px 4px rgba(0,0,0,.1)}.tab-content.active{display:block}.progress-container{width:100%;background-color:#e0e0e0;border-radius:25px;margin:15px 0;position:relative;height:30px}.progress-bar{height:100%;background-color:#2ecc71;border-radius:25px;text-align:center;line-height:30px;color:#fff;width:0;transition:width .5s ease-in-out}.progress-label{position:absolute;width:100%;text-align:center;top:0;left:0;line-height:30px;font-weight:700;color:#333}.subject-progress-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(250px,1fr));gap:20px;margin-top:20px}.progress-card{padding:15px;background:#fff;border-left:5px solid #3498db}.progress-card h3{margin-top:0}
.subject-table-container{display:grid;grid-template-columns:repeat(auto-fit,minmax(900px,1fr));gap:20px}
table{width:100%;border-collapse:collapse;font-size:14px}th,td{padding:10px;border:1px solid #ddd;text-align:left}th{background-color:#ecf0f1}td a{text-decoration:none;color:#2980b9;font-weight:500}td a:hover{text-decoration:underline}.status-done{background-color:#d4edda;color:#155724;font-weight:700}.status-pending{background-color:#fff3cd;color:#856404}.downloads-section{margin-top:40px;padding-top:20px;border-top:2px solid #e0e0e0;text-align:center}.download-link{display:inline-block;margin:5px 10px;padding:10px 18px;background-color:#7f8c8d;color:#fff;text-decoration:none;border-radius:5px;font-weight:700;transition:background-color .3s}.download-link:hover{background-color:#6c7a7d}
.collapsible-header{background-color:#ecf0f1;padding:12px 15px;cursor:pointer;border:1px solid #ddd;border-radius:4px 4px 0 0;margin-top:15px;display:flex;justify-content:space-between;align-items:center}.collapsible-header h3{margin:0;font-size:16px;color:#34495e}.collapsible-header::after{content:'\\25BC';font-size:14px;transition:transform .3s}.collapsible-header.active::after{transform:rotate(180deg)}.collapsible-content{display:none;overflow:hidden;border:1px solid #ddd;border-top:none;border-radius:0 0 4px 4px}.collapsible-content table{margin-bottom:0}
.expandable-image-row td{padding:15px;background-color:#f8f9fa;text-align:center}.expandable-image-row img{max-width:100%;height:auto;border-radius:5px;border:1px solid #ccc}.image-toggle-link{cursor:pointer}
#timer-container{position:fixed;top:20px;right:20px;background-color:#2c3e50;color:white;padding:10px;border-radius:8px;box-shadow:0 4px 8px rgba(0,0,0,0.2);z-index:1000;width:180px}.timer-module{text-align:center;padding:10px;border-bottom:1px solid #34495e}.timer-module:last-child{border-bottom:none}.timer-display{font-size:2em;font-family:monospace;margin-bottom:10px}.timer-buttons{display:flex;gap:8px}.timer-buttons button{flex-grow:1;padding:8px;border:none;border-radius:5px;color:white;font-size:1em;cursor:pointer;transition:background-color .3s}.timer-buttons button:first-child{background-color:#27ae60}.timer-buttons button:first-child:hover{background-color:#229954}.timer-buttons button:first-child.running{background-color:#e74c3c}.timer-buttons button:first-child.running:hover{background-color:#c0392b}.timer-buttons button:last-child{background-color:#7f8c8d}.timer-buttons button:last-child:hover{background-color:#6c7a7d}
.dashboard-layout{display:flex;gap:30px;flex-wrap:wrap}.left-column,.right-column{flex:1;min-width:400px}
.mcq-section{margin-top:0;padding-top:0;border-top:none;}.mcq-section h2{margin-top:0;}.mcq-category{margin-bottom:20px;padding:15px;border-radius:8px;border-left:5px solid}.mcq-category h3{margin-top:0;margin-bottom:15px;border-bottom:1px solid #ddd;padding-bottom:8px}.mcq-link{display:inline-block;margin:4px;padding:8px 15px;color:#fff;text-decoration:none;border-radius:5px;font-weight:500;transition:background-color .3s}
.mcq-physio{border-color:#2ecc71;background-color:#e8f9ef;}.mcq-physio .mcq-link{background-color:#27ae60}.mcq-physio .mcq-link:hover{background-color:#229954}
.mcq-pharm{border-color:#9b59b6;background-color:#f4eef7;}.mcq-pharm .mcq-link{background-color:#8e44ad}.mcq-pharm .mcq-link:hover{background-color:#7d3c98}
.mcq-anat{border-color:#3498db;background-color:#eaf3fa;}.mcq-anat .mcq-link{background-color:#2980b9}.mcq-anat .mcq-link:hover{background-color:#2471a3}
.mcq-path{border-color:#e74c3c;background-color:#fbecec;}.mcq-path .mcq-link{background-color:#c0392b}.mcq-path .mcq-link:hover{background-color:#a93226}
.mcq-random{border-color:#7f8c8d;background-color:#f2f4f4;}.mcq-random .mcq-link{background-color:#707b7c}.mcq-random .mcq-link:hover{background-color:#626b6c}
"""
    with open(OUTPUT_CSS_FILE, 'w', encoding='utf-8') as f: f.write(css_content)
    print(f"✅ Created '{OUTPUT_CSS_FILE}'")

def write_js_file(data):
    """Writes the JS file."""
    js_template = r"""
const subjectData = {data_placeholder};

const timers = {
    '2min': {
        totalSeconds: 120,
        secondsRemaining: 120,
        intervalId: null,
        state: 'stopped',
        displayEl: null,
        controlEl: null,
        resetEl: null
    },
    '10min': {
        totalSeconds: 600,
        secondsRemaining: 600,
        intervalId: null,
        state: 'stopped',
        displayEl: null,
        controlEl: null,
        resetEl: null
    }
};

function formatTime(seconds) {
    const min = Math.floor(seconds / 60);
    const sec = seconds % 60;
    return `${String(min).padStart(2, '0')}:${String(sec).padStart(2, '0')}`;
}

function updateTimerDisplay(key) {
    const timer = timers[key];
    timer.displayEl.textContent = formatTime(timer.secondsRemaining);
}

function controlTimer(key) {
    const timer = timers[key];

    if (timer.state === 'running') {
        clearInterval(timer.intervalId);
        timer.state = 'paused';
        timer.controlEl.textContent = 'Resume';
        timer.controlEl.classList.remove('running');
    } else {
        timer.state = 'running';
        timer.controlEl.textContent = 'Pause';
        timer.controlEl.classList.add('running');
        
        timer.intervalId = setInterval(() => {
            timer.secondsRemaining--;
            updateTimerDisplay(key);
            if (timer.secondsRemaining <= 0) {
                clearInterval(timer.intervalId);
                timer.state = 'stopped';
                timer.controlEl.textContent = 'Start';
                timer.controlEl.classList.remove('running');
                alert(`${key} timer is up!`);
            }
        }, 1000);
    }
}

function resetTimer(key) {
    const timer = timers[key];
    clearInterval(timer.intervalId);
    timer.state = 'stopped';
    timer.secondsRemaining = timer.totalSeconds;
    timer.controlEl.textContent = 'Start';
    timer.controlEl.classList.remove('running');
    updateTimerDisplay(key);
}

function toggleImage(element) {
    const parentRow = element.closest('tr');
    const imageRow = parentRow.nextElementSibling;
    if (imageRow && imageRow.classList.contains('expandable-image-row')) {
        const isVisible = imageRow.style.display === 'table-row';
        imageRow.style.display = isVisible ? 'none' : 'table-row';
        element.classList.toggle('active', !isVisible);
    }
}

function toggleStatus(element) {
    const subject = element.dataset.subject;
    const sectionIndex = parseInt(element.dataset.sectionIndex, 10);
    const itemIndex = parseInt(element.dataset.itemIndex, 10);
    const item = subjectData[subject].data[sectionIndex].items[itemIndex];

    if (item.status.toLowerCase().includes("done")) {
        item.status = "Pending";
    } else {
        item.status = "Done";
    }
    element.textContent = item.status;
    element.className = item.status.toLowerCase().includes("done") ? "status-done" : "status-pending";
    updateProgress();
}

function renderTables() {
    for (const subjectKey in subjectData) {
        const subject = subjectData[subjectKey];
        const container = document.getElementById(subjectKey);

        if (container) {
            let content = `<h2>${subject.title}</h2><div class="subject-table-container">`;
            subject.data.forEach((section, sectionIdx) => {
                content += `<div class="collapsible-wrapper">`;
                content += `<div class="collapsible-header"><h3>${section.header}</h3></div>`;
                content += `<div class="collapsible-content"><table><tbody>`;

                section.items.forEach((item, itemIdx) => {
                    const statusClass = item.status.toLowerCase().includes("done") ? "status-done" : "status-pending";
                    const isImage = item.hyperlink && /\.(jpeg|jpg|gif|png|svg)$/i.test(item.hyperlink);
                    let linkHtml = item.desc;
                    let imageRowHtml = ''; 

                    if(item.hyperlink) {
                        if(isImage) {
                            linkHtml = `<a href="#" class="image-toggle-link" onclick="toggleImage(this); return false;">${item.desc}</a>`;
                            imageRowHtml = `<tr class="expandable-image-row" style="display: none;"><td colspan="2"><img src="${item.hyperlink}" loading="lazy" alt="${item.desc}"></td></tr>`;
                        } else {
                            linkHtml = `<a href="${item.hyperlink}" target="_blank">${item.desc}</a>`;
                        }
                    }

                    content += `<tr>
                        <td>${linkHtml}</td>
                        <td class="${statusClass}" data-subject="${subjectKey}" data-section-index="${sectionIdx}" data-item-index="${itemIdx}" onclick="toggleStatus(this)">
                            ${item.status}
                        </td>
                    </tr>`;
                    
                    content += imageRowHtml;
                });
                content += `</tbody></table></div></div>`;
            });
            content += `</div>`;
            container.innerHTML = content;
        }
    }
}

function updateProgress() {
    let totalDone = 0;
    let totalItems = 0;
    for (const subjectKey in subjectData) {
        const subject = subjectData[subjectKey];
        let subjectDone = 0;
        let subjectItems = 0;
        subject.data.forEach(section => {
            section.items.forEach(item => {
                subjectItems++;
                if (item.status.toLowerCase().includes("done")) {
                    subjectDone++;
                }
            });
        });
        totalDone += subjectDone;
        totalItems += subjectItems;

        const percentage = subjectItems > 0 ? (subjectDone / subjectItems) * 100 : 0;
        const card = document.getElementById(`${subjectKey}-progress-card`);
        if (card) {
            card.style.borderLeftColor = subject.color;
            card.innerHTML = `<h3>${subject.title}</h3>
            <div class="progress-container">
                <div class="progress-bar" style="width:${percentage.toFixed(1)}%;background-color:${subject.color};"></div>
                <span class="progress-label">${subjectDone} / ${subjectItems} (${percentage.toFixed(1)}%)</span>
            </div>`;
        }
    }

    const totalPercentage = totalItems > 0 ? (totalDone / totalItems) * 100 : 0;
    document.getElementById("total-progress-bar").style.width = `${totalPercentage.toFixed(1)}%`;
    document.getElementById("total-progress-label").textContent = `${totalDone} / ${totalItems} (${totalPercentage.toFixed(1)}%)`;
}

document.addEventListener("DOMContentLoaded", () => {
    const tabs = document.querySelectorAll(".tab-btn");
    const tabContents = document.querySelectorAll(".tab-content");

    tabs.forEach(tab => {
        tab.addEventListener("click", () => {
            tabs.forEach(t => t.classList.remove("active"));
            tabContents.forEach(c => c.classList.remove("active"));
            tab.classList.add("active");
            document.getElementById(tab.dataset.tab).classList.add("active");
        });
    });

    renderTables();
    updateProgress();

    document.querySelector('main').addEventListener('click', function(event) {
        const header = event.target.closest('.collapsible-header');
        if (header) {
            header.classList.toggle('active');
            const content = header.nextElementSibling;
            if (content.style.display === "block") {
                content.style.display = "none";
            } else {
                content.style.display = "block";
            }
        }
    });

    for (const key in timers) {
        timers[key].displayEl = document.getElementById(`timer-display-${key}`);
        timers[key].controlEl = document.getElementById(`timer-control-${key}`);
        timers[key].resetEl = document.getElementById(`timer-reset-${key}`);

        timers[key].controlEl.addEventListener('click', () => controlTimer(key));
        timers[key].resetEl.addEventListener('click', () => resetTimer(key));
        
        updateTimerDisplay(key);
    }
});
"""
    json_data = json.dumps(data, indent=4)
    final_js_content = js_template.replace("{data_placeholder}", json_data)

    with open(OUTPUT_SCRIPT_FILE, 'w', encoding='utf-8') as f:
        f.write(final_js_content)
    print(f"✅ Created interactive '{OUTPUT_SCRIPT_FILE}' with your data.")


if __name__ == "__main__":
    print("--- Starting Interactive Webpage Generator ---")
    web_data = generate_web_data()
    if web_data:
        write_html_file()
        write_css_file()
        write_js_file(web_data)
        print("\n🎉 Success! Your interactive webpage files have been created.")
        print("Upload index.html, style.css, script.js, and the 'downloaded_images' folder to your web host.")
    else:
        print("\n❌ Webpage generation failed. Please check for errors above.")
